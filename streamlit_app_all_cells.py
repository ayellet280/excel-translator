
import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font
import tempfile
import os
import googletrans
from googletrans import Translator

st.title("Excel Translator App – תרגום קבצים עם מילון + תרגום אוטומטי")

uploaded_file = st.file_uploader("העלה קובץ Excel לתרגום", type=["xlsx"])
if uploaded_file:
    df = pd.read_excel(uploaded_file)
    st.write("תצוגה מקדימה של הקובץ:")
    st.dataframe(df)

    glossary_file = "glossary.xlsx"
    if not os.path.exists(glossary_file):
        st.error("קובץ glossary.xlsx לא נמצא בריפוזיטורי. ודא שהוא קיים.")
    else:
        glossary_df = pd.read_excel(glossary_file)
        glossary = dict(zip(glossary_df["English"].astype(str).str.strip(), glossary_df["Hebrew"].astype(str).str.strip()))

        translated_df = df.copy()
        fallback_cells = []

        translator = Translator()

        for row_idx in range(df.shape[0]):
            for col_idx in range(df.shape[1]):
                val = df.iat[row_idx, col_idx]
                if pd.isna(val):
                    continue
                val_str = str(val).strip()
                if val_str in glossary:
                    translated_df.iat[row_idx, col_idx] = glossary[val_str]
                else:
                    try:
                        translated = translator.translate(val_str, src='en', dest='he').text
                        translated_df.iat[row_idx, col_idx] = translated
                        fallback_cells.append((row_idx, col_idx))
                    except:
                        translated_df.iat[row_idx, col_idx] = val_str  # Keep as is if translation fails

        # Save to new Excel preserving styles
        temp_dir = tempfile.mkdtemp()
        output_path = os.path.join(temp_dir, "translated.xlsx")
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            translated_df.to_excel(writer, index=False)
            workbook = writer.book
            worksheet = writer.sheets["Sheet1"]

            light_blue = PatternFill(start_color="CCFFFF", end_color="CCFFFF", fill_type="solid")
            bold_font = Font(bold=True)

            for row_idx, col_idx in fallback_cells:
                cell = worksheet.cell(row=row_idx + 2, column=col_idx + 1)  # +2 for header and 0-index
                cell.fill = light_blue
                cell.font = bold_font

        with open(output_path, "rb") as f:
            st.download_button("📥 הורד את הקובץ המתורגם", f, file_name="translated.xlsx")
