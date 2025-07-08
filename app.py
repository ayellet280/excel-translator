
import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
import io

st.set_page_config(page_title="מתרגם קובצי Excel", layout="wide")
st.title("📘 מתרגם קובצי Excel בתחום השילוח הבינלאומי")

@st.cache_resource
def load_glossary():
    try:
        return pd.read_excel("glossary.xlsx")
    except:
        return pd.DataFrame(columns=["English", "Hebrew"])

glossary_df = load_glossary()

def save_glossary(df):
    df.to_excel("glossary.xlsx", index=False)

uploaded_file = st.file_uploader("העלה קובץ Excel לתרגום", type=["xlsx"])
if uploaded_file:
    wb = load_workbook(uploaded_file)
    ws = wb.active

    glossary_dict = {k.strip().lower(): v.strip() for k, v in zip(glossary_df["English"], glossary_df["Hebrew"])}

    custom_fill = PatternFill(start_color="CCFFFF", end_color="CCFFFF", fill_type="solid")
    bold_font = Font(bold=True)

    current_desc_index = None
    current_comment_index = None

    for row in ws.iter_rows(min_row=1):
        headers = [str(cell.value).strip().lower() if cell.value else "" for cell in row]
        if "description" in headers:
            current_desc_index = headers.index("description")
            current_comment_index = headers.index("comment") if "comment" in headers else None
            continue

        if current_desc_index is not None:
            if current_desc_index < len(row):
                cell = row[current_desc_index]
                if cell.value and isinstance(cell.value, str):
                    val = cell.value.strip().lower()
                    if "לבדיקה" not in val and not any("\u0590" <= c <= "\u05EA" for c in val):
                        matched = False
                        for key in glossary_dict:
                            if key in val:
                                cell.value = glossary_dict[key]
                                matched = True
                                break
                        if not matched:
                            cell.value = f"{cell.value} (לבדיקה)"
                            cell.fill = custom_fill
                            cell.font = bold_font
            if current_comment_index is not None and current_comment_index < len(row):
                cell = row[current_comment_index]
                if cell.value and isinstance(cell.value, str):
                    val = cell.value.strip().lower()
                    if "לבדיקה" not in val and not any("\u0590" <= c <= "\u05EA" for c in val):
                        matched = False
                        for key in glossary_dict:
                            if key in val:
                                cell.value = glossary_dict[key]
                                matched = True
                                break
                        if not matched:
                            cell.value = f"{cell.value} (לבדיקה)"
                            cell.fill = custom_fill
                            cell.font = bold_font

    output = io.BytesIO()
    wb.save(output)
    st.success("הקובץ תורגם בהצלחה!")
    st.download_button("📥 הורד את הקובץ המתורגם", output.getvalue(), file_name="translated_file.xlsx")

st.sidebar.header("הוספת מונחים למילון")
eng_term = st.sidebar.text_input("מונח באנגלית")
heb_term = st.sidebar.text_input("תרגום לעברית")
if st.sidebar.button("➕ הוסף למילון"):
    if eng_term and heb_term:
        glossary_df.loc[len(glossary_df)] = [eng_term.strip(), heb_term.strip()]
        save_glossary(glossary_df)
        st.sidebar.success("המונח נוסף בהצלחה!")
    else:
        st.sidebar.warning("נא למלא את שני השדות")

st.sidebar.header("מילון קיים")
st.sidebar.dataframe(glossary_df)
